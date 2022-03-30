"""
weekly Standardlastprofile calculation

"""

from textwrap import indent
from openpyxl import load_workbook

EXTENTION = ".xlsx"
MOTHER_PATH = "//cifs02/RoamingData$/u2110370/Documents/GitHub/Lastprofile/"


def check_file(filename):
    #day_counter = 1
    
    wb = load_workbook(MOTHER_PATH + filename)
    ws = wb.active

    old_date = ws['A3'].value
    for index in range(3,17667):
        i = "A" + str(index)
        z = "B" + str(index)
        if int(ws[z].value) == 0:
            ws[z].value = 0.028538813
        
        #current_date = ws[i].value
        #if current_date.day != old_date.day:
            #day_counter = day_counter + 1
            #old_date = current_date

    wb.save(MOTHER_PATH + filename)
    
    
def excel_addition(file, destination_file):
    weekend_day_counter = 0
    wb = load_workbook(MOTHER_PATH + file)
    ws = wb.active

    wb_des = load_workbook(MOTHER_PATH + destination_file)
    ws_des = wb_des.active

    old_date = ws['A3'].value

    for index in range(3,17667):
        i = "A" + str(index)
        current_date = ws[i].value

        if current_date.weekday() <= 4:
            j = "B" + str(index)
            index_des = ( ((index-2)%96) + 5)
            if index_des == 5:
                index_des = 101
                weekend_day_counter = weekend_day_counter + 1
            z = "C" + str(index_des)
            
            ws_des[z].value = float(ws_des[z].value) + float(ws[j].value)
        
       
    wb_des.save(MOTHER_PATH + destination_file)
    print("Sucessfull", weekend_day_counter)


def excel_average(file):

    wb = load_workbook(MOTHER_PATH +file)
    ws = wb.active
    for index in range(6, 102):
        z = "C" + str(index)
        ws[z].value = float(ws[z].value)  / 132
    
    wb.save(MOTHER_PATH +file)
    print("Average Complet...")


#check_file("SPL_Normallader" + EXTENTION)
#excel_addition("SPL_Normallader" + EXTENTION, "SLP -NLL" + EXTENTION)
excel_average("SLP -NLL" + EXTENTION)












