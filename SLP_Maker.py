"""
weekly Standardlastprofile calculation

"""

import os
from openpyxl import load_workbook

# Constants
EXTENTION = ".xlsx"
MOTHER_PATH = "//cifs02/RoamingData$/u2110370/Documents/GitHub/Lastprofile/"

def get_filenames(folder):
    xlsx_files = os.listdir(MOTHER_PATH + folder)
    print("Getting filenames Complet...")
    return(xlsx_files)


def change_filename(filepath, week):
    new_filepath =  week + "/" +"Lastprofile_" + week + EXTENTION
    filepath = MOTHER_PATH + filepath
    os.rename(filepath, MOTHER_PATH + new_filepath)
    print("Changing filename Complet...")
    return new_filepath

def excel_operation(path1, path2, path3, path4, path5):
    wb1 = load_workbook(MOTHER_PATH + path1)
    ws1 = wb1.active

    wb2 = load_workbook(MOTHER_PATH + path2)
    ws2 = wb2.active

    wb3 = load_workbook(MOTHER_PATH + path3)
    ws3 = wb3.active

    wb4 = load_workbook(MOTHER_PATH + path4)
    ws4 = wb4.active

    wb5 = load_workbook(MOTHER_PATH + path5)
    ws5 = wb5.active

    for index in range(6, 102):
        z = "C" + str(index)
        ws1[z].value = ( float(ws1[z].value) + float(ws2[z].value) + float(ws3[z].value) + float(ws4[z].value) + float(ws5[z].value)   ) / 5

    wb1.save(MOTHER_PATH + path1)
    print("Saving new workbook Complet...")

def main():
    weeks = ["kw_35"]
    #for i in range(5, 51):
     #   weeks.append("kw_" + str(i))

    for week in weeks:
        print("****************")
        print(week + " Start")
        filenames = get_filenames(week)
        filepaths = []

        if (len(filenames) == 5 ):
            filepaths.append( week + "/" + filenames[0])
            filepaths.append( week + "/" + filenames[1])
            filepaths.append( week + "/" + filenames[2])
            filepaths.append( week + "/" + filenames[3])
            filepaths.append( week + "/" + filenames[4])
        
            filepaths[0] = change_filename(filepaths[0], week)
            #ExcelExport_20220329_105216.xlsx
            excel_operation(filepaths[0] , filepaths[1] , filepaths[2] , filepaths[3] , filepaths[4] )
        else:
            print("Number of files incorrect")
            print(filenames)
        
        print(week + "Complet...")
    
    print("All weeks sucessfully completed...")
    


#main()



