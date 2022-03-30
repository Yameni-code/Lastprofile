"""
weekly Standardlastprofile calculation

"""

from openpyxl import load_workbook

# Constants
EXTENTION = ".xlsx"
MOTHER_PATH = "//cifs02/RoamingData$/u2110370/Documents/GitHub/Lastprofile/"
SLP_PATH = MOTHER_PATH + "SLP_schnellLader.xlsx"


def excel_addition(path1):

    wb = load_workbook(SLP_PATH)
    ws = wb.active

    wb1 = load_workbook(MOTHER_PATH + path1)
    ws1 = wb1.active

    for index in range(6, 102):
        z = "C" + str(index)
        i = "D" + str(index)
        ws[i].value = ( float(ws[i].value) + float(ws1[z].value) ) 

    wb.save(SLP_PATH)
    print("Addidion " + path1 +  " Complet...")


def excel_avearge():

    wb = load_workbook(SLP_PATH)
    ws = wb.active
    for index in range(6, 102):
        z = "D" + str(index)
        ws[z].value = float(ws[z].value)  / 46
    
    wb.save(SLP_PATH)
    print("Average Complet...")



def main_SLP():
    weeks = []
    for i in range(5, 51):
        weeks.append("kw_" + str(i))

    for week in weeks:
        print("****************")
        print(week + " Start")
        
        filepath = week + "/" +"Lastprofile_" + week + EXTENTION

        excel_addition(filepath)
        print(week + " End")
    
    excel_avearge()
    print("All weeks sucessfully completed...")
    

main_SLP()
